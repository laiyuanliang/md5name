# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import traceback
import pymysql
import time
from sys import argv

startTime = time.time()
script,sourcefile = argv  

wb = load_workbook(sourcefile)
ws = wb.worksheets[0]
ws.cell(column=3, row=1).value = u"明文姓名"
db = pymysql.connect('localhost','root','Root@mysql01','test')
cursor = db.cursor()
index = 1
try:
    for row in ws.rows:
        if index > 1:
            hash_name = row[1].value
            print(hash_name)
            sql = "SELECT name FROM user WHERE md5_name ='%s'"%(hash_name)
            cursor.execute(sql)
            result = cursor.fetchone()[0]
            ws.cell(row=index, column=3).value = result
        index += 1
        if index%1000 == 0:
            print(index)
				
except Exception as e:
    print(traceback.print_exc())
    print(f"在第{index}行出错")

db.close()
wb.save(sourcefile)

endTime = time.time()
spendTime = int(round(endTime)) - int(round(startTime))
print(f"This script take {spendTime} second.")
